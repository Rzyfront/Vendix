#!/usr/bin/env python3
"""Fill an .xlsx template with normalized product rows, preserving its format.

Two modes:
  --preview N   print the column mapping + N rendered rows and WRITE NOTHING.
                This is the mandatory approval gate (SKILL.md fase 5).
  (default)     write the output workbook.

Rows are a JSON array of dicts keyed by TEMPLATE HEADER text (accent- and
case-insensitive). Unknown keys are reported, never silently dropped.

Usage:
    fill_template.py --template t.xlsx --rows rows.json --preview 3
    fill_template.py --template t.xlsx --rows rows.json --out final.xlsx \
                     [--sheet "Pedido POP"] [--header-row 1] [--clear-existing]
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from copy import copy

import openpyxl


def norm_key(s) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.lower().replace("_", " ")).strip()


def detect_header_row(ws, scan=15) -> int:
    for r in range(1, min(scan, ws.max_row) + 1):
        texty = [c.value for c in ws[r] if isinstance(c.value, str) and c.value.strip()]
        if len(texty) >= 2:
            return r
    return 1


def build_map(ws, header_row: int) -> dict[str, int]:
    m = {}
    for c in ws[header_row]:
        if c.value is not None and str(c.value).strip():
            m[norm_key(c.value)] = c.column
    return m


def last_data_row(ws, header_row: int, cols) -> int:
    """Last row with real content. ws.max_row overcounts styled empty rows."""
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in cols):
            last = r
    return last


def dv_columns(ws) -> list[tuple[object, set[int]]]:
    out = []
    for dv in ws.data_validations.dataValidation:
        cols: set[int] = set()
        for rng in dv.sqref.ranges:
            cols.update(range(rng.min_col, rng.max_col + 1))
        out.append((dv, cols))
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--rows", required=True, help="JSON array of dicts keyed by header")
    ap.add_argument("--out")
    ap.add_argument("--sheet")
    ap.add_argument("--header-row", type=int)
    ap.add_argument("--clear-existing", action="store_true",
                    help="wipe the template's example rows before writing")
    ap.add_argument("--preview", type=int, default=0)
    a = ap.parse_args()

    if not a.preview and not a.out:
        ap.error("--out required unless --preview")

    with open(a.rows, encoding="utf-8") as f:
        rows = json.load(f)
    if not isinstance(rows, list):
        ap.error("--rows must contain a JSON array")

    keep_vba = a.template.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(a.template, keep_vba=keep_vba)
    ws = wb[a.sheet] if a.sheet else wb.worksheets[0]

    hr = a.header_row or detect_header_row(ws)
    colmap = build_map(ws, hr)
    if not colmap:
        raise SystemExit(f"sin encabezados en la fila {hr} de la hoja '{ws.title}'")

    # ---- mapping report -------------------------------------------------
    used: dict[str, int] = {}
    unknown: set[str] = set()
    for r in rows:
        for k, v in r.items():
            nk = norm_key(k)
            if nk not in colmap:
                unknown.add(k)
                continue
            used.setdefault(nk, 0)
            if v not in (None, "") and str(v).strip():
                used[nk] += 1

    header_by_norm = {norm_key(c.value): str(c.value).strip()
                      for c in ws[hr] if c.value is not None and str(c.value).strip()}

    print(f"Hoja: '{ws.title}'  encabezado en fila {hr}  filas a cargar: {len(rows)}")
    print("\nMapeo columna -> cobertura")
    for nk, col in colmap.items():
        letter = openpyxl.utils.get_column_letter(col)
        filled = used.get(nk, 0)
        state = f"{filled}/{len(rows)} con dato" if nk in used else "SIN DATO"
        print(f"  {letter:>3} {header_by_norm[nk]:<32} {state}")
    if unknown:
        print(f"\n  NO MAPEADAS (no existen en la plantilla): {sorted(unknown)}")

    data_cols = list(colmap.values())
    last = last_data_row(ws, hr, data_cols)
    existing = last - hr
    if existing:
        print(f"\n  La plantilla trae {existing} fila(s) de ejemplo bajo el encabezado.")
        print(f"  {'SE BORRARAN' if a.clear_existing else 'SE CONSERVARAN (usar --clear-existing para borrarlas)'}")

    # ---- preview gate ---------------------------------------------------
    if a.preview:
        print(f"\nVista previa de {min(a.preview, len(rows))} registro(s):")
        for r in rows[: a.preview]:
            print("  " + "-" * 60)
            for nk, col in colmap.items():
                val = next((v for k, v in r.items() if norm_key(k) == nk), None)
                shown = "(vacio)" if val in (None, "") else repr(val)
                print(f"    {header_by_norm[nk]:<32} {shown}")
        print("\nNo se escribio ningun archivo. Confirmar el mapeo antes de generar.")
        return 0

    # ---- style donor ----------------------------------------------------
    style_row = hr + 1 if existing else None

    if a.clear_existing and existing:
        for r in range(hr + 1, last + 1):
            for c in data_cols:
                ws.cell(row=r, column=c).value = None
        start = hr + 1
    else:
        start = last + 1

    dvs = dv_columns(ws)

    for i, r in enumerate(rows):
        target = start + i
        for nk, col in colmap.items():
            cell = ws.cell(row=target, column=col)
            if style_row and target != style_row:
                cell._style = copy(ws.cell(row=style_row, column=col)._style)
            val = next((v for k, v in r.items() if norm_key(k) == nk), None)
            cell.value = None if val in ("", None) else val
        # keep dropdowns working on the appended rows
        for dv, cols in dvs:
            for c in cols:
                if c in data_cols:
                    dv.add(ws.cell(row=target, column=c))

    if ws.auto_filter.ref:
        endcol = openpyxl.utils.get_column_letter(max(data_cols))
        ws.auto_filter.ref = f"A{hr}:{endcol}{start + len(rows) - 1}"

    os.makedirs(os.path.dirname(os.path.abspath(a.out)), exist_ok=True)
    wb.save(a.out)
    print(f"\nEscrito {a.out}: {len(rows)} filas en '{ws.title}' desde la fila {start}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
