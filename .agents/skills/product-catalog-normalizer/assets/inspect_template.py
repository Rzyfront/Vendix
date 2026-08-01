#!/usr/bin/env python3
"""Preflight inspection of an .xlsx template before filling it.

Reports what openpyxl WILL preserve and what it WILL DROP, so the agent can warn
the user instead of silently degrading the deliverable.

Usage:
    inspect_template.py <plantilla.xlsx> [--json]
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile

import openpyxl

# openpyxl round-trip drops these parts entirely
LOSSY_PARTS = {
    "xl/drawings/": "imagenes/formas",
    "xl/charts/": "graficos",
    "xl/pivotCache/": "tablas dinamicas",
    "xl/pivotTables/": "tablas dinamicas",
    "xl/threadedComments/": "comentarios en hilo",
    "xl/ctrlProps/": "controles de formulario",
    "xl/activeX/": "controles ActiveX",
}


def detect_header_row(ws, scan=15):
    """First row with >=2 non-empty text cells. Templates rarely deviate."""
    for r in range(1, min(scan, ws.max_row) + 1):
        vals = [c.value for c in ws[r]]
        texty = [v for v in vals if isinstance(v, str) and v.strip()]
        if len(texty) >= 2:
            return r
    return 1


def inspect(path: str) -> dict:
    report: dict = {"file": path, "lossy": [], "sheets": []}

    with zipfile.ZipFile(path) as z:
        names = z.namelist()
    for prefix, label in LOSSY_PARTS.items():
        if any(n.startswith(prefix) for n in names):
            report["lossy"].append(label)
    report["macros"] = any(n.startswith("xl/vbaProject") for n in names)

    wb = openpyxl.load_workbook(path)
    report["defined_names"] = list(wb.defined_names.keys())

    for ws in wb.worksheets:
        hr = detect_header_row(ws)
        headers = [
            (c.column_letter, str(c.value).strip())
            for c in ws[hr]
            if c.value is not None and str(c.value).strip()
        ]
        sample = []
        for row in ws.iter_rows(min_row=hr + 1, max_row=min(hr + 3, ws.max_row), values_only=True):
            if any(v is not None for v in row):
                sample.append(list(row))

        report["sheets"].append(
            {
                "name": ws.title,
                "state": ws.sheet_state,
                "dimensions": ws.dimensions,
                "header_row": hr,
                "headers": headers,
                "sample_rows": sample,
                "sample_row_count": len(sample),
                "freeze_panes": ws.freeze_panes,
                "merged": [str(r) for r in ws.merged_cells.ranges],
                "data_validations": [
                    {"type": dv.type, "formula1": dv.formula1, "sqref": str(dv.sqref)}
                    for dv in ws.data_validations.dataValidation
                ],
                "conditional_formats": len(list(ws.conditional_formatting)),
                "autofilter": str(ws.auto_filter.ref) if ws.auto_filter.ref else None,
                "protected": bool(ws.protection.sheet),
            }
        )
    return report


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("template")
    ap.add_argument("--json", action="store_true")
    a = ap.parse_args()

    rep = inspect(a.template)
    if a.json:
        print(json.dumps(rep, ensure_ascii=False, indent=2, default=str))
        return 0

    print(f"Plantilla: {rep['file']}")
    if rep["lossy"]:
        print(f"  AVISO openpyxl perdera: {', '.join(rep['lossy'])}")
    else:
        print("  Sin graficos/imagenes/pivots -> round-trip seguro con openpyxl")
    if rep["macros"]:
        print("  Contiene macros VBA -> guardar como .xlsm con keep_vba=True")
    for s in rep["sheets"]:
        print(f"\n  Hoja '{s['name']}' ({s['state']}) rango {s['dimensions']}")
        print(f"    fila de encabezado: {s['header_row']}")
        for col, h in s["headers"]:
            print(f"      {col}: {h}")
        print(f"    filas de ejemplo bajo el encabezado: {s['sample_row_count']}")
        for r in s["sample_rows"]:
            print(f"      {r}")
        if s["data_validations"]:
            print("    validaciones:")
            for dv in s["data_validations"]:
                print(f"      {dv['sqref']} {dv['type']} = {dv['formula1']}")
        if s["protected"]:
            print("    hoja PROTEGIDA")
    return 0


if __name__ == "__main__":
    sys.exit(main())
